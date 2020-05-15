# -*- coding:utf-8 -*-
# @Time: 2019/12/12 16:17
# @Author: wenqin_zhu
# @File: component_openpyxl.py
# @Software: PyCharm

"""
openpyxl模块操作Excel
一、进行 - 读
    1、导入 from openpyxl import load_workbook
    2、加载读取指定的workbook
    3、读取指定的表单，也可以通过 .active(),读取当前的活动表单
    4、读取指定表单内的数据.val = ws["表单名称"].value()
    5、打印获取到的表单数据 val
"""


from openpyxl import load_workbook



class ComExcel(object):

    def __init__(self, filename):
        self.filename = filename

    def read_excel(self, sheetname=None):
        #
        try:
            open_wb = load_workbook(self.filename)
        except:
            print(f"读取文件路径错误！{self.filename}")
        else:
            #
            if sheetname is None:
                ws = open_wb.active
            else:
                try:
                    ws = open_wb[sheetname]
                except:
                    print(f"读取表单错误！{open_wb[sheetname]}")
            # 遍历读取所有表单
            # return tuple(ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1, values_only=False))[0]
            # list_val = []
            # for val in list_val:
            #     list_val.append(val)
            # return list_val

            head_data_tuple = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
            one_list = []
            for one_tuple in tuple(ws.iter_rows(min_row=2, values_only=True)):
                one_list.append(dict(zip(head_data_tuple, one_tuple)))

            return one_list


if __name__ == '__main__':
    from tools.common_os import CASE_DATAS
    """openpyxl操作Excel步骤"""
    # 1、加载worlbook，并打开指定的Excel文件
    wb = load_workbook(CASE_DATAS)
    # 2、打开指定的worksheet
    ws = wb["login"]
    # 3、读取指定表单内的数据
    print(ws["A1"].value)
    a = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
    print(a)

    # val = ComExcel(CASE_DATAS).read_excel()
    # print(val)
